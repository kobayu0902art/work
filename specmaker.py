import openpyxl as px

combo=r"path"
edit=r"path"
source=r"path"
alloc=r"path"
goalpath=r"path"
basestr="sample"

cwb=px.load_workbook(combo)
cws=cwb.active

ewb=px.load_workbook(edit)
ews=ewb.active

swb=px.load_workbook(source)
sws=swb.active

awb=px.load_workbook(alloc)
aws=awb.active

combolist=(83,10,8,7,9)#6~
editlist=(147,84,85,85,87)#84~

#specmake(84,ewb,ews)
def specmake(start,wb,ws,path):
    for j in range(5):
        #print(editlist[j])
        for i in range(start,editlist[j]+1):
            wb=px.load_workbook(edit)
            ws=wb.active
            #print(swb.sheetnames)
            sws=swb[swb.sheetnames[j]]
            casenumber=str(sws.cell(row=i,column=1).value)
            print(f"現在{casenumber}処理中")
            i_one=basestr+casenumber
            ws.title=casenumber
            finalpath=goalpath+"\\"+path+"\\"+casenumber+".xlsx"
            ws.cell(row=1,column=9,value=i_one)
            c_six_base=sws.cell(row=i,column=2).value
            tempc_six=str(sws.cell(row=i,column=3).value).splitlines()
            l=[(6, 3), (7, 3), (8, 3), (9, 3), (10, 3), (11, 3), (12, 3), (13, 3), (6, 18), (7, 18), (8, 18), (9, 18), (10, 18), (11, 18), (12, 18), (13, 18)]
            if len(tempc_six)>=2:
                for k in range(len(tempc_six)):
                    c_six=c_six_base+"-"+tempc_six[k]
                    #print(c_six)
                    #print("l[k]",l[k])
                    ws.cell(row=l[k][0],column=l[k][1],value=c_six)
            else:
                ws.cell(row=6,column=3,value=c_six_base)
                #print(ws.cell(row=6,column=3,value=c_six_base).value)
            wb.save(finalpath)
            print("保存完了")

specmake(6,cwb,cws,"Combobox")

specmake(84,ewb,ews,"Editbox")

