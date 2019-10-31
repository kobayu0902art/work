import os
import glob
import openpyxl as px
import shutil

#優先探索path
primary=[
    r"path"
]
#次に探索するpath
secondary=[
    r"path",
    r"path",
    r"path",
    r"path",
    r"path",
    r"path"
]
destination=r"path"
specdest=destination+"\\"+"テスト仕様書"
eviddest=destination+"\\"+"Evidence"

pl=len(primary)
sl=len(secondary)

#上記pathを新規ファイルのpathとして使えるように加工
for q in range(pl):
    primary[q]+="\\"
    #print(primary[q])
    
for q in range(sl):
    secondary[q]+="\\"
    #print(secondary[q])

xls=".xls"

#20191016 xlsで固定とし、ハードコーディングしてしまう

def search(path,filename):
    if path == primary:
        len = pl
    elif path == secondary:
        len = sl
    for i in range(len):
            os.chdir(path[i])
            tempfilename=path[i]+"**\\"+filename+xls
            l=glob.glob(tempfilename, recursive=True)
            if bool(l)==True:
                break
            #新旧全部漏れなく記載する場合はl→templにしてif bool(l)==true:l.append(templ)【l=[]の宣言必要】
    return l
#ex)search(primary,casename) search(secondary,casename)

def evidsearch(path,filename):
    if path == primary:
        len = pl
    elif path == secondary:
        len = sl
    for i in range(len):
            os.chdir(path[i])
            tempfilename=path[i]+"**\\"+filename#+os.sep
            l=glob.glob(tempfilename, recursive=True)
            if bool(l)==True and bool(os.listdir(l[0]))==True:
                break
    return l

#ex)copy(pathlist[0],specdest)
#pathフォルダの中身をコピーする
#try-exceptionを簡単に書きたい
def filecopy(origin,destination):
    try:
        shutil.copy(origin,destination)
    #親フォルダの中に子フォルダがあるとフォルダごとコピーしようとするが、
    #フォルダごとの移動でpermissionerrorを起こしてしまうため、errorを検知したら同名フォルダを作成し、中身をコピー
    #子フォルダの中に孫フォルダがあると同じエラーが起こるため、3階層分ほどtryしてみる
    #もっとスマートな書き方があるはず
    except PermissionError:
        folder=destination+"\\"+casename
        try:
            os.mkdir(folder)
        except FileExistsError:
            pass
        for i in os.listdir(origin):
            idvdest=folder+"\\"+i
            idvorig=origin+"\\"+i
            try:
                shutil.copy(idvorig,idvdest)
            except PermissionError:
                for j in os.listdir(idvorig):
                    idvdest=idvdest+"\\"+j
                    idvorig=idvorig+"\\"+j
                try:
                    shutil.copy(idvorig,idvdest)
                except PermissionError:
                    for k in os.listdir(idvorig):
                        idvdest=idvdest+"\\"+k
                        idvorig=idvorig+"\\"+k
                    try:
                        shutil.copy(idvorig,idvdest)
                    except PermissionError:
                        for l in os.listdir(idvorig):
                            idvdest=idvdest+"\\"+l
                            idvorig=idvorig+"\\"+l
                        try:
                            shutil.copy(idvorig,idvdest)
                        except FileNotFoundError:
                            pass
                except FileNotFoundError:
                    pass
    except FileExistsError:
        pass
    print("コピー完了")

workbook=r"path"
wb=px.load_workbook(workbook)
caselist=wb['caselist']
tup=((4,4387),(5,53),(6,53),(7,12),(12,129),(13,365),(14,12))
for count in range(len(tup)):
    print(f"現在{tup[count][0]}番検索中")
    end=tup[count][1]
    resultname="result_"+str(tup[count][0])
    result=wb[resultname]
    result.cell(row=1,column=1,value="case-number")
    result.cell(row=1,column=2,value="仕様書path")
    result.cell(row=1,column=3,value="フォルダpath")
    result.cell(row=1,column=4,value="フォルダ中身")
    for i in range(2,end+2):
        if i % 50==0:
            wb.save(workbook)
        casename=caselist.cell(row=i,column=count+1).value
        print(f"進捗率:{i-1}/{end}")

        print("sampleを検索します(仕様書)")
        pathlist=search(primary,casename)
        if bool(pathlist)==True:
            result.cell(row=i,column=1,value=casename)
            result.cell(row=i,column=2,value=str(pathlist[0]))
            filecopy(pathlist[0],specdest)
        elif bool(pathlist)==False:
            print("さらに過去も検索")
            pathlist=search(secondary,casename)
            if bool(pathlist)==True:
                result.cell(row=i,column=1,value=casename)
                result.cell(row=i,column=2,value=str(pathlist[0]))
                filecopy(pathlist[0],specdest)
            elif bool(pathlist)==False:
                result.cell(row=i,column=1,value=casename)
                result.cell(row=i,column=2,value="該当なし")
                print(f"\n{casename}:該当なし\n")

        print("sampleを検索します")
        evidlist=evidsearch(primary,casename)
        if bool(evidlist)==True:
            result.cell(row=i,column=3,value=str(evidlist[0]))
            result.cell(row=i,column=4,value=str(os.listdir(evidlist[0])))
            filecopy(evidlist[0],eviddest)
        elif bool(evidlist)==False:
            print("さらに過去も検索")
            evidlist=evidsearch(secondary,casename)
            if bool(evidlist)==True:
                result.cell(row=i,column=3,value=str(evidlist[0]))
                result.cell(row=i,column=4,value=str(os.listdir(evidlist[0])))
                filecopy(evidlist[0],eviddest)
            elif bool(evidlist)==False:
                result.cell(row=i,column=3,value="該当なし")
                print(f"\n{casename}:該当なし\n")
    count+=1
    wb.save(workbook)
    
wb.save(workbook)
input("End")