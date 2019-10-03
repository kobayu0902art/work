#!/usr/bin/env python
# coding: utf-8

#sampleに置換済み

import os
import pandas as pd
import openpyxl as px
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Color, PatternFill
import unicodedata

#日本語検知
def is_japanese(string):
    for ch in string:
        name = unicodedata.name(ch) 
        if "CJK UNIFIED" in name \
        or "HIRAGANA" in name \
        or "KATAKANA" in name:
            return True
    return False

#仕様書フォルダ
spec=r"sample"
#エビデンスフォルダ
evid=r"sample"
bs="\\"
xls=".xls"
spec+=bs
evid+=bs

#チェック対象一覧ブック
#A2からタテに並ぶ想定
assignbook=r"sample"
#チェック結果出力ブック
checkbook=r"sample"

#wb,wsに一覧ブック格納、出力ブックにコピー、出力ブック作成
wb = px.load_workbook(assignbook)
wb.save(checkbook)
ws=wb.active

#項目名記入
ws.cell(row=1,column=2,value="日付")
ws.cell(row=1,column=3,value="仕様書エビデンスファイル数")
ws.cell(row=1,column=4,value="実際のエビデンスファイル数")
ws.cell(row=1,column=5,value="CD TrueFalse")
ws.cell(row=1,column=6,value="仕様書エビデンスファイル名")
ws.cell(row=1,column=7,value="実際のエビデンスファイル名")

#一覧ブックをDataFrameとして読み込み、リスト化、項目数をlに格納
worklist=pd.read_excel(assignbook)
worklist=worklist.values.tolist()
l=len(worklist)

#項目数ぶん処理
for i in range(l):
    #階層ごとのフォルダ名であることを利用し、フルパス作成
    one=str(worklist[i])[2:9]
    two=str(worklist[i])[2:12]
    three=str(worklist[i])[2:16]
    finalspecpath=spec+one+bs+two+bs+three+xls
    finalevidpath=evid+one+bs+two+bs+three
    #エビデンスファイル一覧
    filelist=os.listdir(finalevidpath)
    fllen=len(filelist)
    #仕様書をDataFrameとして読み込む
    specwb=pd.read_excel(finalspecpath, header=None) 
    #そのなかから大体この範囲にファイルパスが書かれているであろう範囲だけ切り取る
    evidarea=specwb.loc[20:60,2:40]
    #その範囲だけ一時ブックとして作成(後で削除)
    evidarea.to_excel(r"sample")
    tempwb=px.load_workbook(r"sample")
    tempws=tempwb.active
    #エビデンスファイルのファイルパスのリスト作成(1項目に無駄データ作成しているが後で削除)
    evidlist=list(range(1))
    #タテ方向走査(ファイル名の順番をそろえるため)
    for col_num in range(1,40):
        for row_num in range(1,42):
            #tempエクセルの空欄を埋める(これをしないと次の処理エラー)
            if tempws.cell(row=row_num,column=col_num).value == None:
                tempws.cell(row=row_num,column=col_num,value="temp")
            #ファイルパスに日本語が含まれているか判定(基本はファイルパスに日本語は含まれず、テスト手順の誤検知を防ぐため(改善の余地あり))
            jap=is_japanese(str(tempws.cell(row=row_num,column=col_num).value))
            #sampleが含まれるand日本語が含まれない=ファイルパスとし、リストに格納
            if "sample" in str(tempws.cell(row=row_num,column=col_num).value) and jap == False:
                evidlist.append(tempws.cell(row=row_num,column=col_num).value)
    #evidlistの無駄データ削除
    del evidlist[0]
    #仕様書記載のエビデンスファイル名のリスト作成
    samplelen=len(evidlist)
    #仕様書記載の日付
    ws.cell(row=i+2,column=2,value=str(specwb.at[1,38]))
    #仕様書記載エビデンスファイル数
    ws.cell(row=i+2,column=3,value=samplelen)
    #エビデンスフォルダ内のエビデンスファイル数
    ws.cell(row=i+2,column=4,value=fllen)

    ipt=str(i+2)
    #仕様書と実際のエビデンスファイル数があっているか判定
    tf="=C"+ipt+"=D"+ipt
    ws.cell(row=i+2,column=5,value=tf)
    #仕様書記載のエビデンスファイル名
    ws.cell(row=i+2,column=6,value=str(evidlist))
    #エビデンスフォルダ内のエビデンスファイル名
    ws.cell(row=i+2,column=7,value=str(filelist))
    
#E列のFalse強調表示
range="E2:E"+ipt
ws.conditional_formatting.add(range,
                              CellIsRule(operator='equal',formula=['FALSE'], 
                                         fill=PatternFill(start_color='FF0000', end_color='FF0000',fill_type='solid')))
#一時ブック削除
os.remove(r"sample")
#出力ブック保存
wb.save(checkbook)
