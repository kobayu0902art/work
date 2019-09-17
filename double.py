import openpyxl as px
import datetime


now=datetime.datetime.now()
now=str(now)
now=now.replace(":"," ")

obj=input("元ファイル\n")
sr,sc,er,ec=[int(i) for i in input("sr,sc,er,ec\n").split()]


wb=px.load_workbook(obj)
#39 1 52 1  39 1 66 4
abc=input("sheet\n")
ws=wb[abc]
resultb=px.load_workbook(r"sample")
results=resultb.create_sheet(now)
diff=-1
for i in range(sr,er+1):
    diff+=1
    for j in range(sc,ec+1):
        temp=ws.cell(row=i,column=j).value
        results.cell(row=i+diff,column=j,value=temp)
        results.cell(row=i+diff+1,column=j,value=temp)

resultb.save(r"sample")