#ライブラリ、モジュールをインポート
import pandas as pd
import openpyxl as px
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Color, PatternFill

#ブック名入力
tdname=input('testdataName?')
edname=input('editordataName?')

#読み込んだブックの同じテストのデータをDataFrameに格納
td=pd.read_excel(tdname,header=1,sheet_name=0)
ed=pd.read_excel(edname,header=1,sheet_name=0)

#テストデータのラベルの部分をリストに変換
tdlabel=td.columns.values
tdlabel=tdlabel.tolist()

#テストデータのラベルの長さを格納(後のループ処理(rangeの引数)に使用)
l=len(tdlabel)

#DataFrameの定義
#定義不要かもしれない
add=pd.DataFrame()
result=pd.DataFrame()

#エディタデータ成形
#テストデータラベルの項目名と正規表現でマッチする部分のエディタデータを検索、成形
for i in range(l):
    add = ed.loc[:,ed.columns.str.match(tdlabel[i])]
    result = pd.concat([result,add],axis=1)

#TrueFalse判定
tf=pd.DataFrame()
tf=td==result

#出力ブックの名前指定
outname=input('outputName?')

#各データを1ブック２シートに出力
with pd.ExcelWriter(outname) as writer:
    tf.to_excel(writer,sheet_name='TrueFalse')
    td.to_excel(writer,sheet_name='TestData')
    result.to_excel(writer,sheet_name='EditorData')
    
#FALSEを強調表示する条件付き書式設定
wb=px.load_workbook(outname)
ws=wb['TrueFalse']
ws.conditional_formatting.add('A1:AZ100',CellIsRule(operator='equal',formula=['FALSE'],
                                                    fill=PatternFill(start_color='FF0000', end_color='FF0000',
                                                    fill_type='solid')))
white=px.styles.PatternFill(patternType='solid',
                            fgColor='000000', bgColor='000000')
ws['A1'].fill=white
ws.conditional_formatting.add('A1:AZ100',CellIsRule(operator='equal',formula=[''],
                                                    fill=PatternFill(start_color='000000', end_color='000000',
                                                    fill_type='solid')))
wb.save(outname)
