import openpyxl as px
import xlrd
import os
from datetime import datetime, timedelta

def rabel_loc(j):
    for i in range(ws.nrows):
        if ws.cell(i,1).value==rabel[j]:
            return i
        
def excel_date(num):
    return(datetime(1899, 12, 30) + timedelta(days=num))

def area_copy():
    for i in range(1,29):
        for j in range(len(xlsl)):
            templates.cell(row=xlsxl[j],column=i+1,value=str(ws.cell(xlsl[j],i).value))

srcdir=r"path"
template=r"path"
savedir=r"path"
osl=os.listdir(r"path")

rabel=['sample', 'sample', 'sample', 'sample', 'sample']

#1はじまり
#i1,ac2,af2,aj2,am2
headertup=((9,1), (29,2), (32,2), (36,2), (39,2))
#B6AD9 B11AL16 B18AP23 B25AP45 B48AP75
#contenttup=( ((2,6),(30,9)) , ((2,11),(38,16)) , ((2,18),(39,23)) , ((2,25),(39,45)))# , ((2,48),(39,75)) )
contenttup=( (6,9) , (11,16) , (18,23) , (25,45) , (48,75) )

for k in osl:
    print(f"現在{k}処理中")
    templateb=px.load_workbook(template)
    templates=templateb.active
    path=srcdir+"\\"+k
    wb=xlrd.open_workbook(path)
    ws=wb.sheet_by_index(0)

    #header
    for i in range(len(headertup)):
        output=(ws.cell(headertup[i][1]-1,headertup[i][0]-1).value)
        if type(ws.cell(headertup[i][1]-1,headertup[i][0]-1).value) == float:
            output=excel_date(ws.cell(headertup[i][1]-1,headertup[i][0]-1).value)
            output=format(output,'%y/%m/%d')
            output="20"+output
        #print(output)
        templates.cell(row=headertup[i][1],column=headertup[i][0],value=str(output).replace("sample","sample"))



    for j in range(len(rabel)):
        if j == 4:
            xlsl=[_ for _ in range(rabel_loc(j)+1,ws.nrows)]
            xlsxl=[_ for _ in range(contenttup[j][0],contenttup[j][0]+len(xlsl))]
            try:
                area_copy()
            except AttributeError:#結合セルがread_onlyのため結合セルに対してws.cellでエラー
                print(f"{k}でエラー発生")
                pass
            break
        xlsl=[_ for _ in range(rabel_loc(j)+1,rabel_loc(j+1))]
        xlsxl=[_ for _ in range(contenttup[j][0],contenttup[j][0]+len(xlsl))]
        try:
            area_copy()
        except AttributeError:
            print(f"{k}でエラー発生")
            pass
    savepath=savedir+"\\"+k+"x"
    savepath=savepath.replace("sample","sample")
    print(f"{savepath}:保存済み")
    templateb.save(savepath)