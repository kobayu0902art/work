import openpyxl as px

#I1 C6 9 1  3 6
#row0col8 番号  row5col2 小項目
combotemplate=r"path"
edittemplate=r"path"
srcpath=r"path"
goalpath=r"path"
base=r"path"
combotempb=px.load_workbook(combotemplate)
combotemps=combotempb.active
edittempb=px.load_workbook(edittemplate)
edittemps=edittempb.active
srcb=px.load_workbook(srcpath)
l=srcb.sheetnames
ll=[78,5,3,2,4,64,1,2,2,4]
for i in range(10):
    print("i:",i)
    for j in range(6,ll[i]+6):
        if i<5:
            print("j:",j)
            srcs=srcb[l[i]]
            print(l[i])
            print(str(srcs.cell(row=j,column=1).value))
            i_one=base+str(srcs.cell(row=j,column=1).value)
            #rowとcol逆ぅ
            combotemps.cell(row=9,column=1,value=i_one)
            tempc_six=str(srcs.cell(row=j,column=3).value).splitlines()
            for k in range(len(tempc_six)):
                print("k:",k,"k+6:",k+6)
                c_six=str(srcs.cell(row=j,column=2).value)+tempc_six[k]
                if k+6<10:
                    combotemps.cell(row=k+6,column=3,value=c_six)
                elif k+6>=10:
                    print("2行目")
                    combotemps.cell(row=k+6,column=19,value=c_six)
            finalpath=goalpath+"\\"+str(srcs.cell(row=j,column=1).value)+".xlsx"
            combotempb.save(finalpath)
        else:
            print("j+ll[i-5]:",j+ll[i-5])
            srcs=srcb[l[i-5]]
            print(l[i-5])
            print(str(srcs.cell(row=j+ll[i-5],column=1).value))
            i_one=base+str(srcs.cell(row=j+ll[i-5],column=1).value)
            edittemps.cell(row=9,column=1,value=i_one)
            tempc_six=str(srcs.cell(row=j+ll[i-5],column=3).value).splitlines()
            for k in range(len(tempc_six)):
                print("k:",k,"k+6:",k+6)
                c_six=str(srcs.cell(row=j+ll[i-5],column=2).value)+tempc_six[k]
                if k+6<10:
                    edittemps.cell(row=k+6,column=3,value=c_six)
                elif k+6>=10:
                    print("2行目")
                    edittemps.cell(row=k+6,column=19,value=c_six)
            finalpath=goalpath+"\\"+str(srcs.cell(row=j+5+ll[i-5],column=1).value)+".xlsx"
            edittempb.save(finalpath)