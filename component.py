import openpyxl as px
import pandas as pd
import datetime


now=datetime.datetime.now()
now=str(now)
now=now.replace(":"," ")
read=r"sample"
component=r"sample"


start,end= [int(i) for i in input().split()]
area=range(start,end+1)

wb=px.load_workbook(read)
ws=wb["テストパターン"]
cb=px.load_workbook(component)
cs=cb.create_sheet(now)
cb.save(component)
l=[]
for w in area:
    for i in range(9,18):
        ii=i+1
        if ws.cell(row=w,column=ii).value=="○":
            l.append(ws.cell(row=3,column=ii).value)
    s=""
    if len(l)!=0:
        for q in range(len(l)):
            s+=l[q]
            s+="\n"
            cs.cell(row=w,column=1,value=s)
    else:
        cs.cell(row=w,column=1,value="該当なし")
    l=[]
cb.save(component)