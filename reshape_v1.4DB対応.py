#ライブラリ、モジュールをインポート
import pandas as pd
import openpyxl as px
from openpyxl.formatting.rule import CellIsRule
from openpyxl.styles import Color, PatternFill
from openpyxl.utils import get_column_letter

#ブック名入力
tdname=input('testdataName?')
edname=input('editordataName?')
ddname=input('DBdataName?')

#読み込んだブックの同じテストのデータをDataFrameに格納
td=pd.read_excel(tdname,header=1,sheet_name=0)
ed=pd.read_excel(edname,header=1,sheet_name=0)
dd=pd.read_excel(ddname,header=1,sheet_name=0)

#テストデータのラベルの部分をリストに変換
tdlabel=td.columns.values
tdlabel=tdlabel.tolist()

#テストデータのラベルの長さを格納(後のループ処理(rangeの引数)に使用)
l=len(tdlabel)

#DataFrameの定義
#定義不要かもしれない
add=pd.DataFrame()
result=pd.DataFrame()
dbadd=pd.DataFrame()
dbresult=pd.DataFrame()

#エディタデータ成形
#テストデータラベルの項目名と正規表現でマッチする部分のエディタデータを検索、成形
for i in range(l):
    dbadd = dd.loc[:,dd.columns.str.match(tdlabel[i])]
    dbresult = pd.concat([dbresult,dbadd],axis=1)

#エディタデータ成形
#テストデータラベルの項目名と正規表現でマッチする部分のエディタデータを検索、成形
for i in range(l):
    add = ed.loc[:,ed.columns.str.match(tdlabel[i])]
    result = pd.concat([result,add],axis=1)

#TrueFalse判定
tf=pd.DataFrame()
tf=td==result

tdbd=pd.DataFrame()
tdbd=td==dbresult

eddb=pd.DataFrame()
eddb=dbresult==result

#出力ブックの名前指定
outname=input('outputName?')

#各データを1ブック3シートに出力
with pd.ExcelWriter(outname) as writer:
    tf.to_excel(writer,sheet_name='TrueFalse')
    td.to_excel(writer,sheet_name='TestData')
    result.to_excel(writer,sheet_name='EditorData')
    
#FALSEを強調表示する条件付き書式設定
wb=px.load_workbook(outname)
ws=wb['TrueFalse']
colletter = get_column_letter(l)
tdindex=td.index.values
tdindex=tdindex.tolist()
vl=len(tdindex)+1
targetrange = 'B2'+ ':' + colletter + str(vl)
ws.conditional_formatting.add(targetrange,CellIsRule(operator='equal',formula=['FALSE'], 
                                                    fill=PatternFill(start_color='FF0000', end_color='FF0000',
                                                    fill_type='solid')))
wb.save(outname)
